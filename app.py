import streamlit as st
import pandas as pd
import bcrypt
import re
from supabase import create_client
from io import BytesIO
from openpyxl.styles import Alignment

st.set_page_config(page_title="BBJ Reviu", layout="wide")

# =========================
# SUPABASE CONNECTION
# =========================
url = st.secrets["supabase"]["url"]
key = st.secrets["supabase"]["key"]
supabase = create_client(url, key)


@st.cache_data(ttl=60)
def get_siap(dinas):
    return supabase.table("neraca_siap") \
        .select("kode_rekening,nama_rekening,saldo_akhir") \
        .eq("dinas", dinas).execute().data

@st.cache_data(ttl=60)
def get_sipd(dinas):
    return supabase.table("neraca_sipd") \
        .select("kode_rekening,nama_rekening,saldo_akhir") \
        .eq("dinas", dinas).execute().data

# =========================
# LOGIN FUNCTION
# =========================
def cek_login(username, password):
    res = supabase.table("user_login")\
        .select("*")\
        .eq("username", username)\
        .eq("is_active", True)\
        .execute()

    if res.data:
        user = res.data[0]
        if bcrypt.checkpw(password.encode(), user["password_hash"].encode()):
            return True, user["nama_staf"]
    return False, None

# =========================
# LOGIN UI
# =========================
if "login" not in st.session_state:
    st.session_state.login = False

if not st.session_state.login:
    st.title("🔐 Login BBJ (Online)")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        valid, nama = cek_login(username, password)
        if valid:
            st.session_state.login = True
            st.session_state.nama = nama
            st.rerun()
        else:
            st.error("Username / Password salah")
    st.stop()

# =========================
# SIDEBAR
# =========================
st.sidebar.write(f"👤 {st.session_state.nama}")
if st.sidebar.button("Logout"):
    st.session_state.login = False
    st.rerun()

# =========================
# SESSION INIT
# =========================
defaults = {
    "load_dinas": False,
    "dinas_terakhir": None,
    "sudah_simpan_siap": False,
    "mode_revisi_siap": False,
    "trigger_revisi_siap": 0,
    "mode_revisi_sipd": False,
    "trigger_revisi_sipd": 0,
    "hitung_selisih": False,
    "df_merge": None,
    "boleh_simpan": False,
    "sudah_simpan_jurnal": False,
}
for k,v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# =========================
# LIST DINAS (SINGKATKAN JIKA MAU)
# =========================
# =========================
# LIST DINAS
# =========================
list_dinas = [
    "SMK NEGERI 1 SURABAYA",
    "SMK NEGERI 5 SURABAYA",
    "SMK NEGERI 6 SURABAYA",
    "SMK NEGERI 1 BUDURAN SIDOARJO",
    "SMK NEGERI 3 BUDURAN SIDOARJO",
    "SMK NEGERI 2 MALANG",
    "SMK NEGERI 4 MALANG",
    "SMK NEGERI 11 MALANG",
    "SMK NEGERI 1 SINGOSARI MALANG",
    "SMK NEGERI 1 PANJI SITUBONDO",
    "SMK NEGERI 1 KALIPURO BANYUWANGI",
    "SMK NEGERI 2 BONDOWOSO",
    "SMK NEGERI 5 JEMBER",
    "SMK NEGERI 3 MADIUN",
    "SMK NEGERI 1 PACITAN",
    "SMK NEGERI 2 KOTA PASURUAN",
    "SMK NEGERI 3 BOYOLANGU TULUNGAGUNG",
    "SMK NEGERI 1 GLAGAH BANYUWANGI",
    "SMK NEGERI 1 TEGALAMPEL BONDOWOSO",
    "SMK NEGERI 1 JENANGAN PONOROGO",
    "SMK NEGERI 2 NGANJUK",
    "SMK NEGERI 1 BANYUWANGI",
    "SMK NEGERI 1 LUMAJANG",
    "SMK NEGERI KALIBARU BANYUWANGI",
    "SMK NEGERI 6 JEMBER",
    "SMK NEGERI DARUL ULUM BANYUWANGI",
    "SMK NEGERI 1 PASURUAN",
    "SMK NEGERI 2 TUBAN",
    "SMK NEGERI 1 GRATI PASURUAN",
    "SMK NEGERI 1 PUNGGING MOJOKERTO",
    "SMK NEGERI 10 SURABAYA",
    "SMK NEGERI 2 BATU",
    "SMK NEGERI 2 JIWAN MADIUN",
    "SMK NEGERI 2 PROBOLINGGO",
    "SMK NEGERI 7 SURABAYA",
    "SMK NEGERI 1 BENDO MAGETAN",
    "SMK NEGERI 1 NGANJUK",
    "SMK NEGERI 1 GEMPOL PASURUAN",
    "SMK NEGERI RENGEL TUBAN",

    "RUMAH SAKIT UMUM DAERAH dr. SOETOMO",
    "RUMAH SAKIT UMUM DAERAH dr. SAIFUL ANWAR",
    "RUMAH SAKIT UMUM DAERAH dr. SOEDONO MADIUN",
    "RUMAH SAKIT UMUM DAERAH HAJI PROVINSI JAWA TIMUR",
    "RUMAH SAKIT JIWA MENUR",
    "RUMAH SAKIT UMUM DAERAH KARSA HUSADA BATU",
    "RUMAH SAKIT PARU JEMBER",
    "RUMAH SAKIT UMUM DAERAH DUNGUS",
    "RUMAH SAKIT UMUM DAERAH DAHA HUSADA",
    "RUMAH SAKIT UMUM DAERAH SUMBERGLAGAH",
    "RUMAH SAKIT MATA MASYARAKAT JAWA TIMUR",
    "RUMAH SAKIT UMUM DAERAH HUSADA PRIMA",
    "RUMAH SAKIT UMUM DAERAH MOHAMMAD NOER PAMEKASAN",
    "RUMAH SAKIT PARU MANGUHARJO PROVINSI JAWA TIMUR",
    "UPT PELATIHAN KESEHATAN MASYARAKAT MURNAJATI",

    "UPT PENGEMBANGAN BENIH PADI DAN PALAWIJA",
    "UPT PENGEMBANGAN BENIH HORTIKULTURA",
    "UPT PENGEMBANGAN AGRIBISNIS TANAMAN PANGAN DAN HORTIKULTURA",

    "UPT PELABUHAN PERIKANAN PANTAI MAYANGAN",
    "UPT PELABUHAN PERIKANAN PANTAI TAMPERAN",
    "UPT PELABUHAN PERIKANAN PANTAI PONDOKDADAP"
]

# =========================
# HELPER
# =========================
def extract_nama_dinas(text):
    return re.sub(r"^\s*\(.*?\)\s*", "", str(text)).strip()

def normalisasi_nama(nama):
    return nama.upper().replace(".", "").strip()

def cocokkan_dinas(nama_excel):
    nama_excel_norm = normalisasi_nama(nama_excel)
    for d in list_dinas:
        if normalisasi_nama(d) in nama_excel_norm:
            return d
    return None

def format_rupiah(angka):
    return f"{angka:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# =========================
# UI
# =========================
st.title("📊 Upload Neraca Saldo BLUD")

with st.expander("📂 Pilih Dinas BLUD", expanded=True):

    dinas = st.selectbox("Pilih Dinas", list_dinas)

    if st.session_state.dinas_terakhir != dinas:
        st.session_state.load_dinas = False
        st.session_state.dinas_terakhir = dinas

    if st.button("🔍 Load"):
        st.session_state.load_dinas = True
        st.session_state.dinas = dinas
        st.session_state.hitung_selisih = False 
        st.session_state.sudah_simpan_jurnal = False
        st.rerun()

# =========================
# MAIN
# =========================
if st.session_state.load_dinas:

    st.success(f"✅ Dinas: {st.session_state.dinas}")
    col1, col2 = st.columns(2)

    # =========================
    # SIAP
    # =========================
    # =========================
    # SIAP (PERBAIKAN FINAL)
    # =========================
    # =========================
    # SIAP (FIX TOMBOL REVISI)
    # =========================
    with col1:
        data_siap_db = get_siap(st.session_state.dinas)
        df_siap_db = pd.DataFrame(data_siap_db)
        
        total_db = df_siap_db["saldo_akhir"].astype(float).sum() if not df_siap_db.empty else 0
        
        # Update status simpan berdasarkan data DB
        st.session_state.sudah_simpan_siap = total_db > 0

        st.subheader("📥 Neraca SIAP")

        # LOGIKA TERBALIK: Cek Mode Revisi DULU, baru cek status DB
        if st.session_state.mode_revisi_siap or not st.session_state.sudah_simpan_siap:
            # TAMPILKAN UPLOADER
            file = st.file_uploader("Upload SIAP", type=["xlsx"], key=f"siap_{st.session_state.trigger_revisi_siap}")

            if file:
                df = pd.read_excel(file, header=None, dtype=str)
                dinas_file = extract_nama_dinas(df.iloc[5,4])
                match = cocokkan_dinas(dinas_file)

                if match != st.session_state.dinas:
                    st.error(f"❌ Dinas tidak sesuai. File milik: {dinas_file}")
                    st.stop()

                # --- PROSES DATA ---
                data = df.iloc[7:].copy()[[1,2,3,4,8]]
                data.columns = ["kode","u1","u2","u3","saldo"]
                data_8102 = data[data["kode"].astype(str).str.startswith("8102")].copy()
                data_8102["nama"] = (data_8102["u1"].fillna("")+" "+data_8102["u2"].fillna("")+" "+data_8102["u3"].fillna("")).str.strip()
                data_8102["saldo"] = pd.to_numeric(data_8102["saldo"].str.replace(",", ""), errors="coerce").fillna(0)

                total_excel = data_8102["saldo"].sum()
                st.success(f"✅ Data Siap. Total: Rp {format_rupiah(total_excel)}")

                if st.button("💾 Simpan SIAP ke Database"):
                    supabase.table("neraca_siap").delete().eq("dinas", match).execute()
                    insert_siap = [
                        {
                            "dinas": match,
                            "kode_rekening": r["kode"],
                            "nama_rekening": r["nama"],
                            "saldo_akhir": float(r["saldo"]),
                            "is_active": True
                        } for _, r in data_8102.iterrows()
                    ]
                    if insert_siap:
                        for i in range(0, len(insert_siap), 500):
                            supabase.table("neraca_siap").insert(insert_siap[i:i+500]).execute()
                    
                    get_siap.clear() 
                    st.session_state.mode_revisi_siap = False # MATIKAN MODE REVISI SETELAH SIMPAN
                    st.success("✅ Berhasil simpan!")
                    st.rerun()
            
            # Tombol Batal jika sedang dalam mode revisi
            if st.session_state.mode_revisi_siap and total_db > 0:
                if st.button("❌ Batal Revisi"):
                    st.session_state.mode_revisi_siap = False
                    st.rerun()

        else:
            # TAMPILKAN INFO DB & TOMBOL UPLOAD ULANG
            st.info(f"Total di Database: Rp {format_rupiah(total_db)}")
            if st.button("🔄 Upload Ulang SIAP"):
                st.session_state.mode_revisi_siap = True
                st.session_state.trigger_revisi_siap += 1
                st.session_state.hitung_selisih = False
                st.rerun()
    # =========================
    # SIPD (FINAL VERSION)
    # =========================
    # =========================
    # SIPD (WITH BATAL REVISI)
    # =========================
    with col2:
        # 1. Ambil data yang sudah ada di database
        data_sipd = get_sipd(st.session_state.dinas)
        df_sipd_db = pd.DataFrame(data_sipd)
        
        total_sipd_db = df_sipd_db["saldo_akhir"].astype(float).sum() if not df_sipd_db.empty else 0
        st.session_state.sudah_simpan_sipd = total_sipd_db > 0

        # Ambil total SIAP untuk pembanding balance
        res_siap = supabase.table("neraca_siap").select("saldo_akhir").eq("dinas", st.session_state.dinas).execute()
        total_siap = sum(float(x["saldo_akhir"]) for x in res_siap.data) if res_siap.data else 0

        st.subheader("📥 Neraca SIPD")

        # Tampilan Form Upload (Jika Belum Ada Data ATAU Sedang Mode Revisi)
        if st.session_state.mode_revisi_sipd or not st.session_state.sudah_simpan_sipd:
            file = st.file_uploader("Upload File Excel SIPD", type=["xlsx"], key=f"sipd_{st.session_state.trigger_revisi_sipd}")

            if file:
                df = pd.read_excel(file, header=None, dtype=str)
                dinas_file = extract_nama_dinas(df.iloc[2,2])
                match = cocokkan_dinas(dinas_file)

                if match != st.session_state.dinas:
                    st.error(f"❌ File milik: {dinas_file}. Harus sesuai pilihan!")
                    st.stop()

                # Proses data
                data = df.iloc[7:].copy()[[0,1,8,9]]
                data.columns = ["kode","nama","debit","kredit"]
                data["kode"] = data["kode"].str.replace(".","", regex=False)

                def clean_money(x):
                    if pd.isna(x): return 0
                    return pd.to_numeric(str(x).replace(".","").replace(",","."), errors="coerce")

                data["debit"] = data["debit"].apply(clean_money).fillna(0)
                data["kredit"] = data["kredit"].apply(clean_money).fillna(0)
                data["saldo"] = data["debit"] - data["kredit"]
                
                df_final_sipd = data[data["kode"].str.startswith("8102")].copy()
                total_excel = df_final_sipd["saldo"].sum()

                # Validasi Balance
                selisih_cek = abs(total_excel - total_siap)
                if selisih_cek > 1:
                    st.error(f"""
                    ❌ Tidak balance dengan SIAP!
                    
                    **Total SIAP** : Rp {format_rupiah(total_siap)}
                    
                    **Total SIPD** : Rp {format_rupiah(total_excel)}
                    
                    **Selisih** : Rp {format_rupiah(total_excel - total_siap)}
                    """)
                    st.stop()
                
                st.success(f"✅ Balance! Total SIPD: Rp {format_rupiah(total_excel)}")

                if st.button("💾 Simpan SIPD ke Database"):
                    supabase.table("neraca_sipd").delete().eq("dinas", match).execute()
                    data_insert = [
                        {
                            "dinas": match,
                            "kode_rekening": r["kode"],
                            "nama_rekening": r["nama"],
                            "saldo_akhir": float(r["saldo"]),
                            "is_active": True
                        } for _, r in df_final_sipd.iterrows()
                    ]
                    for i in range(0, len(data_insert), 500):
                        supabase.table("neraca_sipd").insert(data_insert[i:i+500]).execute()

                    get_sipd.clear()
                    st.session_state.mode_revisi_sipd = False
                    st.success("✅ Data berhasil disimpan!")
                    st.rerun()

            # --- TOMBOL BATAL REVISI (Hanya muncul jika di DB sudah ada data) ---
            if st.session_state.mode_revisi_sipd and total_sipd_db > 0:
                if st.button("❌ Batal Revisi", key="btn_batal_sipd"):
                    st.session_state.mode_revisi_sipd = False
                    st.rerun()

        # Tampilan jika sudah ada data di database
        else:
            st.info(f"Total di Database: Rp {format_rupiah(total_sipd_db)}")

            if st.button("🔄 Upload Ulang SIPD"):
                st.session_state.mode_revisi_sipd = True
                st.session_state.trigger_revisi_sipd += 1
                st.rerun()

                         
# ==================================
# 1. TOMBOL PEMICU (HITUNG SELISIH)
# ==================================
if st.session_state.sudah_simpan_siap and st.session_state.sudah_simpan_sipd:
    st.markdown("---")
    if st.button("🔍 Hitung Selisih SIAP vs SIPD", use_container_width=True):
        st.session_state.hitung_selisih = True
        st.session_state.boleh_simpan = True
        st.session_state.sudah_simpan_jurnal = False
        st.rerun()

# ==================================
# 2. BLOK LOGIKA & TAMPILAN TABEL
# ==================================
if st.session_state.get("hitung_selisih"):
    
    # Ambil data terbaru dari DB
    data_siap_db = get_siap(st.session_state.dinas)
    data_sipd_db = get_sipd(st.session_state.dinas)
    
    df1 = pd.DataFrame(data_siap_db)
    df2 = pd.DataFrame(data_sipd_db)

    # --- Sinkronisasi Kolom SIAP ---
    if df1.empty:
        df1 = pd.DataFrame(columns=["kode_rekening", "nama_rekening_siap", "siap"])
    else:
        df1 = df1.rename(columns={"saldo_akhir": "siap", "nama_rekening": "nama_rekening_siap"})

    # --- Sinkronisasi Kolom SIPD ---
    if df2.empty:
        df2 = pd.DataFrame(columns=["kode_rekening", "nama_rekening_sipd", "sipd"])
    else:
        df2 = df2.rename(columns={"saldo_akhir": "sipd", "nama_rekening": "nama_rekening_sipd"})

    # Pastikan numerik
    df1["siap"] = pd.to_numeric(df1["siap"], errors="coerce").fillna(0)
    df2["sipd"] = pd.to_numeric(df2["sipd"], errors="coerce").fillna(0)

    # MERGE & KALKULASI
    df = pd.merge(df1, df2, on="kode_rekening", how="outer")
    df["nama_rekening"] = df["nama_rekening_siap"].combine_first(df["nama_rekening_sipd"])
    df["siap"] = df["siap"].fillna(0)
    df["sipd"] = df["sipd"].fillna(0)
    df["selisih"] = df["siap"] - df["sipd"]

    # Filter & Reorder Kolom
    df_final = df[["kode_rekening", "nama_rekening", "siap", "sipd", "selisih"]].copy()
    st.session_state.df_merge = df_final.copy() # Simpan data asli untuk jurnal

    # --- DISPLAY TABEL ---
    st.subheader("📊 Perbandingan SIAP vs SIPD")
    
    df_display = df_final.copy()
    for col in ["siap", "sipd", "selisih"]:
        df_display[col] = df_display[col].apply(format_rupiah)
    
    st.dataframe(df_display, use_container_width=True)

    # Total Selisih
    total_selisih = df_final["selisih"].sum()
    st.markdown(f"""
        <div style="margin-top:10px; padding:15px; border-radius:10px; background-color:#f0f2f6; text-align:right;">
            <span style="font-size:16px;">Total Selisih:</span><br>
            <span style="font-size:24px; font-weight:bold;">Rp {format_rupiah(total_selisih)}</span>
        </div>
    """, unsafe_allow_html=True)

    # ==================================
    # 3. TOMBOL SIMPAN JURNAL
    # ==================================
    if st.session_state.boleh_simpan:
        st.write("")
        if st.button("💾 Simpan Jurnal Ke Database"):
            # Hapus lama
            supabase.table("hasil_perbandingan").delete().eq("dinas", st.session_state.dinas).execute()
            
            # Siapkan data (Hanya yang ada selisihnya)
            df_filtered = df_final[df_final["selisih"] != 0]
            
            if not df_filtered.empty:
                insert_data = df_filtered.apply(lambda r: {
                    "nomor_bukti": "JP Reviu Inspektorat/BBJ BLUD/2025",
                    "tanggal_bukti": "2025-12-31",
                    "keterangan": f"Jurnal Rinci BBJ BLUD {st.session_state.dinas}",
                    "kode_bas": r["kode_rekening"],
                    "uraian": r["nama_rekening"],
                    "debit": r["selisih"] if r["selisih"] > 0 else 0,
                    "kredit": abs(r["selisih"]) if r["selisih"] < 0 else 0,
                    "keterangan_rinci": "-",
                    "dinas": st.session_state.dinas,
                    "is_active": True
                }, axis=1).tolist()
                
                supabase.table("hasil_perbandingan").insert(insert_data).execute()
                st.session_state.sudah_simpan_jurnal = True
                st.success("✅ Jurnal berhasil disimpan!")
                import time
                time.sleep(1.5) 
                st.rerun()
            else:
                st.warning("Tidak ada selisih yang perlu dibuatkan jurnal.")

    # ==================================
    # 4. EXPORT EXCEL (FORMAT ANGKA FIX)
    # ==================================
    if st.session_state.sudah_simpan_jurnal:
        res = supabase.table("hasil_perbandingan") \
                .select("nomor_bukti,tanggal_bukti,keterangan,kode_bas,uraian,debit,kredit,keterangan_rinci") \
                .eq("dinas", st.session_state.dinas) \
                .order("kode_bas").execute()
        
        if res.data:
            df_export = pd.DataFrame(res.data)
            
            # Rapikan Nama Kolom
            df_export.columns = [
                "Nomor Bukti", "Tanggal Bukti", "Keterangan", 
                "Kode BAS", "Uraian", "Debit", "Kredit", "Keterangan Rinci"
            ]
            
            # Kosongkan baris duplikat untuk kolom A, B, C (biar cantik)
            df_export_final = df_export.copy()
            if len(df_export_final) > 1:
                df_export_final.loc[1:, ["Nomor Bukti", "Tanggal Bukti", "Keterangan"]] = ""

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export_final.to_excel(writer, index=False, sheet_name='jurnal')
                ws = writer.sheets['jurnal']
                
                # --- PENGATURAN FORMAT EXCEL ---
                
                # 1. Lebar Kolom
                widths = {'A': 40, 'B': 15, 'C': 50, 'D': 18, 'E': 45, 'F': 20, 'G': 20, 'H': 25}
                for col, width in widths.items():
                    ws.column_dimensions[col].width = width
                
                # 2. Format Angka (Kolom F & G)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    # Kolom Debit (Cell F)
                    cell_debit = ws.cell(row=row_idx, column=6)
                    val_debit = float(cell_debit.value) if cell_debit.value else 0
                    cell_debit.value = val_debit
                    cell_debit.number_format = '0.00' # Di file aslinya 1000000.50, di Excel Indo jadi 1000000,50
                    
                    # Kolom Kredit (Cell G)
                    cell_kredit = ws.cell(row=row_idx, column=7)
                    val_kredit = float(cell_kredit.value) if cell_kredit.value else 0
                    cell_kredit.value = val_kredit
                    cell_kredit.number_format = '0.00'

                    # Alignment tetap rapi di atas
                    for cell in row:
                        cell.alignment = Alignment(vertical="top")

            processed_data = output.getvalue()

            st.download_button(
                label="📥 Download Excel Jurnal",
                data=processed_data,
                file_name=f"Jurnal_{st.session_state.dinas}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )






