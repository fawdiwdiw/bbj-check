from sqlalchemy import create_engine, text
import streamlit as st
import bcrypt
import pandas as pd
import re
from io import BytesIO

st.set_page_config(page_title="BBJ Reviu", layout="wide")

# =========================
# POSTGRES CONNECTION
# =========================
DB_USER = "postgres"
DB_PASS = "postgres"
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "bbj_db"

engine = create_engine(
    f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
)

# =========================
# LOGIN FUNCTION
# =========================
def cek_login(username, password):
    with engine.begin() as conn:
        user = conn.execute(text("""
            SELECT username, password_hash, nama_staf
            FROM user_login
            WHERE username=:u AND is_active=TRUE
        """), {"u": username}).fetchone()

    if user:
        if bcrypt.checkpw(password.encode(), user.password_hash.encode()):
            return True, user.nama_staf
    return False, None

# =========================
# LOGIN UI
# =========================
if "login" not in st.session_state:
    st.session_state.login = False

if not st.session_state.login:
    st.title("🔐 Login BBJ")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        valid, nama = cek_login(username, password)

        if valid:
            st.session_state.login = True
            st.session_state.nama = nama
            st.rerun()
        else:
            st.error("Username / Password salah")

    st.stop()

# =========================
# SIDEBAR
# =========================
st.sidebar.write(f"👤 {st.session_state.nama}")

if st.sidebar.button("Logout"):
    st.session_state.login = False
    st.rerun()

# =========================
# SESSION INIT
# =========================
if "load_dinas" not in st.session_state:
    st.session_state.load_dinas = False

if "dinas_terakhir" not in st.session_state:
    st.session_state.dinas_terakhir = None

if "sudah_simpan_siap" not in st.session_state:
    st.session_state.sudah_simpan_siap = False

if "mode_revisi_siap" not in st.session_state:
    st.session_state.mode_revisi_siap = False

if "last_uploaded_siap" not in st.session_state:
    st.session_state.last_uploaded_siap = None

if "trigger_revisi_siap" not in st.session_state:
    st.session_state.trigger_revisi_siap = 0

if "df_merge" not in st.session_state:
    st.session_state.df_merge = None

if "hitung_selisih" not in st.session_state:
    st.session_state.hitung_selisih = False

# =========================
# LIST DINAS
# =========================
list_dinas = [
    "SMK NEGERI 1 SURABAYA",
    "SMK NEGERI 5 SURABAYA",
    "SMK NEGERI 6 SURABAYA",
    "SMK NEGERI 1 BUDURAN SIDOARJO",
    "SMK NEGERI 3 BUDURAN SIDOARJO",
    "SMK NEGERI 2 MALANG",
    "SMK NEGERI 4 MALANG",
    "SMK NEGERI 11 MALANG",
    "SMK NEGERI 1 SINGOSARI MALANG",
    "SMK NEGERI 1 PANJI SITUBONDO",
    "SMK NEGERI 1 KALIPURO BANYUWANGI",
    "SMK NEGERI 2 BONDOWOSO",
    "SMK NEGERI 5 JEMBER",
    "SMK NEGERI 3 MADIUN",
    "SMK NEGERI 1 PACITAN",
    "SMK NEGERI 2 KOTA PASURUAN",
    "SMK NEGERI 3 BOYOLANGU TULUNGAGUNG",
    "SMK NEGERI 1 GLAGAH BANYUWANGI",
    "SMK NEGERI 1 TEGALAMPEL BONDOWOSO",
    "SMK NEGERI 1 JENANGAN PONOROGO",
    "SMK NEGERI 2 NGANJUK",
    "SMK NEGERI 1 BANYUWANGI",
    "SMK NEGERI 1 LUMAJANG",
    "SMK NEGERI KALIBARU BANYUWANGI",
    "SMK NEGERI 6 JEMBER",
    "SMK NEGERI DARUL ULUM BANYUWANGI",
    "SMK NEGERI 1 PASURUAN",
    "SMK NEGERI 2 TUBAN",
    "SMK NEGERI 1 GRATI PASURUAN",
    "SMK NEGERI 1 PUNGGING MOJOKERTO",
    "SMK NEGERI 10 SURABAYA",
    "SMK NEGERI 2 BATU",
    "SMK NEGERI 2 JIWAN MADIUN",
    "SMK NEGERI 2 PROBOLINGGO",
    "SMK NEGERI 7 SURABAYA",
    "SMK NEGERI 1 BENDO MAGETAN",
    "SMK NEGERI 1 NGANJUK",
    "SMK NEGERI 1 GEMPOL PASURUAN",
    "SMK NEGERI RENGEL TUBAN",

    "RUMAH SAKIT UMUM DAERAH dr. SOETOMO",
    "RUMAH SAKIT UMUM DAERAH dr. SAIFUL ANWAR",
    "RUMAH SAKIT UMUM DAERAH dr. SOEDONO MADIUN",
    "RUMAH SAKIT UMUM DAERAH HAJI PROVINSI JAWA TIMUR",
    "RUMAH SAKIT JIWA MENUR",
    "RUMAH SAKIT UMUM DAERAH KARSA HUSADA BATU",
    "RUMAH SAKIT PARU JEMBER",
    "RUMAH SAKIT UMUM DAERAH DUNGUS",
    "RUMAH SAKIT UMUM DAERAH DAHA HUSADA",
    "RUMAH SAKIT UMUM DAERAH SUMBERGLAGAH",
    "RUMAH SAKIT MATA MASYARAKAT JAWA TIMUR",
    "RUMAH SAKIT UMUM DAERAH HUSADA PRIMA",
    "RUMAH SAKIT UMUM DAERAH MOHAMMAD NOER PAMEKASAN",
    "RUMAH SAKIT PARU MANGUHARJO PROVINSI JAWA TIMUR",
    "UPT PELATIHAN KESEHATAN MASYARAKAT MURNAJATI",

    "UPT PENGEMBANGAN BENIH PADI DAN PALAWIJA",
    "UPT PENGEMBANGAN BENIH HORTIKULTURA",
    "UPT PENGEMBANGAN AGRIBISNIS TANAMAN PANGAN DAN HORTIKULTURA",

    "UPT PELABUHAN PERIKANAN PANTAI MAYANGAN",
    "UPT PELABUHAN PERIKANAN PANTAI TAMPERAN",
    "UPT PELABUHAN PERIKANAN PANTAI PONDOKDADAP"
]

# =========================
# HELPER FUNCTION
# =========================
def extract_nama_dinas(text):
    return re.sub(r"^\s*\(.*?\)\s*", "", str(text)).strip()

def normalisasi_nama(nama):
    return nama.upper().replace(".", "").strip()

def cocokkan_dinas(nama_excel):
    nama_excel_norm = normalisasi_nama(nama_excel)
    for d in list_dinas:
        if normalisasi_nama(d) in nama_excel_norm:
            return d
    return None

def format_rupiah(angka):
    return f"{angka:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# =========================
# UI
# =========================
st.title("📊 Upload Neraca Saldo BLUD")

with st.expander("📂 Pilih Dinas BLUD", expanded=True):

    dinas = st.selectbox("Pilih Dinas", list_dinas)

    if st.session_state.dinas_terakhir != dinas:
        st.session_state.load_dinas = False
        st.session_state.dinas_terakhir = dinas

    if st.button("🔍 Load"):
        st.session_state.load_dinas = True
        st.session_state.dinas = dinas

# =========================
# UPLOAD
# =========================
if st.session_state.load_dinas:

    st.success(f"✅ Dinas terpilih: {st.session_state.dinas}")

    col1, col2 = st.columns(2)

    # =========================
    # SIAP
    # =========================
    with col1:

        # =========================
        # SESSION TAMBAHAN
        # =========================
        if "trigger_revisi_siap" not in st.session_state:
            st.session_state.trigger_revisi_siap = 0

        # =========================
        # CEK DATA DI DB (AUTO LOAD)
        # =========================
        with engine.begin() as conn:
            cek_data = conn.execute(text("""
                SELECT SUM(saldo_akhir) as total
                FROM neraca_siap
                WHERE dinas = :d
            """), {"d": st.session_state.dinas}).fetchone()

        if cek_data and cek_data.total:
            st.session_state.sudah_simpan_siap = True
            total_db = cek_data.total
        else:
            st.session_state.sudah_simpan_siap = False
            total_db = 0

        # =========================
        # HEADER
        # =========================
        if st.session_state.sudah_simpan_siap:
            st.subheader("📥 Upload Neraca SIAP ✅")
        else:
            st.subheader("📥 Upload Neraca SIAP")

        # =========================
        # MODE: SUDAH TERSIMPAN
        # =========================
        if st.session_state.sudah_simpan_siap and not st.session_state.mode_revisi_siap:

            st.markdown(
                f"""
                <div style="padding:12px; border-radius:12px; background-color:#e6f4ea;">
                    <b>✅ Merupakan Neraca Saldo SIAP pada {st.session_state.dinas}</b><br>
                    <span style="font-size:26px; font-weight:bold;">
                        Rp {format_rupiah(total_db)}
                    </span>
                </div>
                """,
                unsafe_allow_html=True
            )

            # tombol upload ulang
            if st.button("🔄 Upload Ulang"):
                st.session_state.mode_revisi_siap = True
                st.session_state.trigger_revisi_siap += 1
                st.rerun()

        # =========================
        # MODE: UPLOAD / REVISI
        # =========================
        else:

            file_siap = st.file_uploader(
                "Upload Excel SIAP",
                type=["xlsx"],
                key=f"siap_{st.session_state.trigger_revisi_siap}"
            )

            if file_siap:
                try:
                    df = pd.read_excel(file_siap, header=None, dtype=str)

                    # =========================
                    # AMBIL DINAS
                    # =========================
                    dinas_raw = df.iloc[5, 4]
                    dinas_clean = extract_nama_dinas(dinas_raw)
                    dinas_match = cocokkan_dinas(dinas_clean)

                    st.info(f"📄 Dinas di file: {dinas_clean}")

                    if dinas_match == st.session_state.dinas:

                        # =========================
                        # AMBIL DATA
                        # =========================
                        data = df.iloc[7:].copy()
                        data = data[[1,2,3,4,8]]
                        data.columns = ["kode","u1","u2","u3","saldo"]

                        data["kode"] = data["kode"].astype(str).str.strip()

                        # filter 8102
                        data_8102 = data[data["kode"].str.startswith("8102")].copy()

                        # gabung nama
                        data_8102["nama"] = (
                            data_8102["u1"].fillna('') + " " +
                            data_8102["u2"].fillna('') + " " +
                            data_8102["u3"].fillna('')
                        ).str.strip()

                        # bersihkan saldo
                        data_8102["saldo"] = data_8102["saldo"].str.replace(",", "", regex=True)
                        data_8102["saldo"] = pd.to_numeric(data_8102["saldo"], errors="coerce").fillna(0)

                        # =========================
                        # TOTAL
                        # =========================
                        total_saldo = data_8102["saldo"].sum()

                        st.markdown(
                            f"""
                            <div style="padding:10px; border-radius:10px; background-color:#e6f4ea;">
                                <b>✅ Merupakan Neraca Saldo SIAP pada {dinas_match}</b><br>
                                <span style="font-size:24px; font-weight:bold;">
                                    Rp {format_rupiah(total_saldo)}
                                </span>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        # =========================
                        # CEK BBJ
                        # =========================
                        cek_bbj = data_8102[data_8102["kode"] == "810299999999"]

                        if not cek_bbj.empty and cek_bbj["saldo"].sum() != 0:

                            nilai = cek_bbj["saldo"].sum()

                            st.error(
                                f"❌ Masih terdapat saldo akhir atas BBJ BLUD senilai Rp {format_rupiah(nilai)}. Silahkan revisi dahulu"
                            )
                            st.stop()

                        else:
                            st.success("✅ Tidak ada saldo akhir atas BBJ BLUD")

                            # =========================
                            # SIMPAN
                            # =========================
                            if st.button("💾 Simpan ke Database"):

                                with engine.begin() as conn:

                                    # hapus semua data lama
                                    conn.execute(text("""
                                        DELETE FROM neraca_siap
                                        WHERE dinas = :d
                                    """), {"d": dinas_match})

                                    # insert baru
                                    for _, row in data_8102.iterrows():
                                        conn.execute(text("""
                                            INSERT INTO neraca_siap
                                            (dinas, kode_rekening, nama_rekening, saldo_akhir, is_active, created_at)
                                            VALUES (:d, :k, :n, :s, TRUE, NOW())
                                        """), {
                                            "d": dinas_match,
                                            "k": row["kode"],
                                            "n": row["nama"],
                                            "s": row["saldo"]
                                        })

                                st.session_state.mode_revisi_siap = False
                                st.session_state.sudah_simpan_siap = True
                                st.rerun()

                    else:
                        st.warning("⚠️ Dinas tidak sesuai, silakan upload ulang")

                except Exception as e:
                    st.error(f"Error: {e}")
    # =========================
    # SIPD
    # =========================
    with col2:

        # =========================
        # SESSION TAMBAHAN
        # =========================
        if "trigger_revisi_sipd" not in st.session_state:
            st.session_state.trigger_revisi_sipd = 0

        if "mode_revisi_sipd" not in st.session_state:
            st.session_state.mode_revisi_sipd = False

        # =========================
        # CEK DATA DI DB (AUTO LOAD)
        # =========================
        with engine.begin() as conn:
            cek_data = conn.execute(text("""
                SELECT SUM(saldo_akhir) as total
                FROM neraca_sipd
                WHERE dinas = :d
            """), {"d": st.session_state.dinas}).fetchone()

        if cek_data and cek_data.total:
            sudah_simpan_sipd = True
            total_db = cek_data.total
        else:
            sudah_simpan_sipd = False
            total_db = 0

        # =========================
        # AMBIL TOTAL SIAP (UNTUK VALIDASI)
        # =========================
        with engine.begin() as conn:
            total_siap_db = conn.execute(text("""
                SELECT COALESCE(SUM(saldo_akhir),0)
                FROM neraca_siap
                WHERE dinas = :d
            """), {"d": st.session_state.dinas}).scalar()

            # FIX: paksa jadi float
            total_siap_db = float(total_siap_db or 0)

        # =========================
        # HEADER
        # =========================
        if sudah_simpan_sipd:
            st.subheader("📥 Upload Neraca SIPD ✅")
        else:
            st.subheader("📥 Upload Neraca SIPD")

        # =========================
        # MODE: SUDAH TERSIMPAN
        # =========================
        if sudah_simpan_sipd and not st.session_state.mode_revisi_sipd:

            st.markdown(
                f"""
                <div style="padding:12px; border-radius:12px; background-color:#e6f4ea;">
                    <b>✅ Neraca SIPD pada {st.session_state.dinas}</b><br>
                    <span style="font-size:26px; font-weight:bold;">
                        Rp {format_rupiah(total_db)}
                    </span>
                </div>
                """,
                unsafe_allow_html=True
            )

            if st.button("🔄 Upload Ulang SIPD"):
                st.session_state.mode_revisi_sipd = True
                st.session_state.trigger_revisi_sipd += 1
                st.rerun()

            # ✅ PINDAH KE SINI (POSISI BENAR)
            if st.session_state.sudah_simpan_siap:
                if st.button("🔍 Hitung Selisih SIAP vs SIPD"):
                    st.session_state.hitung_selisih = True    

        # =========================
        # MODE: UPLOAD / REVISI
        # =========================
        else:

            file_sipd = st.file_uploader(
                "Upload Excel SIPD",
                type=["xlsx"],
                key=f"sipd_{st.session_state.trigger_revisi_sipd}"
            )

            if file_sipd:
                try:
                    df = pd.read_excel(file_sipd, header=None, dtype=str)

                    # =========================
                    # AMBIL DINAS (C BARIS 3)
                    # =========================
                    dinas_raw = df.iloc[2, 2]
                    dinas_clean = extract_nama_dinas(dinas_raw)
                    dinas_match = cocokkan_dinas(dinas_clean)

                    st.info(f"📄 Dinas di file: {dinas_clean}")

                    if dinas_match == st.session_state.dinas:

                        # =========================
                        # AMBIL DATA
                        # =========================
                        data = df.iloc[7:].copy()
                        data = data[[0,1,8,9]]  # A,B,I,J
                        data.columns = ["kode","nama","debit","kredit"]

                        # =========================
                        # FORMAT KODE
                        # =========================
                        data["kode"] = (
                            data["kode"]
                            .astype(str)
                            .str.replace(".", "", regex=False)
                            .str.strip()
                        )

                        # =========================
                        # FILTER 8102 SAJA
                        # =========================
                        data_8102 = data[data["kode"].str.startswith("8102")].copy()

                        # =========================
                        # BERSIHKAN ANGKA (WAJIB UNTUK SIPD)
                        # =========================
                        def clean_angka(x):
                            if pd.isna(x):
                                return 0
                            x = str(x)
                            x = x.replace(".", "")   # hapus ribuan
                            x = x.replace(",", ".")  # koma jadi desimal
                            x = x.replace(" ", "")
                            return pd.to_numeric(x, errors="coerce")

                        data_8102["debit"] = data_8102["debit"].apply(clean_angka).fillna(0)
                        data_8102["kredit"] = data_8102["kredit"].apply(clean_angka).fillna(0)


                        data_8102["saldo"] = data_8102["debit"] - data_8102["kredit"]

                        # =========================
                        # TOTAL
                        # =========================
                        total_saldo = data_8102["saldo"].sum()

                        # =========================
                        # VALIDASI TOTAL SIAP vs SIPD
                        # =========================
                        if total_saldo != total_siap_db:

                            selisih_total = total_saldo - total_siap_db

                            st.error(f"""
                            ❌ Total SIPD tidak sama dengan SIAP  
                            SIAP : Rp {format_rupiah(total_siap_db)}  
                            SIPD : Rp {format_rupiah(total_saldo)}  
                            Selisih : Rp {format_rupiah(selisih_total)}  

                            👉 Silakan revisi terlebih dahulu
                            """)

                            st.stop()
                        else:
                            st.success("✅ Total SIPD sudah sama dengan SIAP")

                        st.markdown(
                            f"""
                            <div style="padding:10px; border-radius:10px; background-color:#e6f4ea;">
                                <b>✅ Neraca SIPD pada {dinas_match}</b><br>
                                <span style="font-size:24px; font-weight:bold;">
                                    Rp {format_rupiah(total_saldo)}
                                </span>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

                        st.success("✅ Data SIPD siap disimpan")

                        # =========================
                        # SIMPAN
                        # =========================
                        if st.button("💾 Simpan SIPD ke Database"):

                            with engine.begin() as conn:

                                # hapus lama
                                conn.execute(text("""
                                    DELETE FROM neraca_sipd
                                    WHERE dinas = :d
                                """), {"d": dinas_match})

                                # insert baru
                                for _, row in data_8102.iterrows():
                                    conn.execute(text("""
                                        INSERT INTO neraca_sipd
                                        (dinas, kode_rekening, nama_rekening, saldo_akhir, is_active, created_at)
                                        VALUES (:d, :k, :n, :s, TRUE, NOW())
                                    """), {
                                        "d": dinas_match,
                                        "k": row["kode"],
                                        "n": row["nama"],
                                        "s": row["saldo"]
                                    })

                            st.session_state.mode_revisi_sipd = False
                            st.rerun()

                    else:
                        st.warning("⚠️ Dinas tidak sesuai, silakan upload ulang")

                except Exception as e:
                    st.error(f"Error: {e}")

        #if sudah_simpan_sipd and st.session_state.sudah_simpan_siap:
            #if st.button("🔍 Hitung Selisih SIAP vs SIPD"):
                #st.session_state.hitung_selisih = True

        # =========================
        # HITUNG & TAMPILKAN SELISIH
        # =========================
        if (
            st.session_state.get("hitung_selisih", False)
            and sudah_simpan_sipd
            and st.session_state.sudah_simpan_siap
        ):

            with engine.begin() as conn:

                df_siap = pd.read_sql("""
                    SELECT kode_rekening, nama_rekening, saldo_akhir
                    FROM neraca_siap
                    WHERE dinas = %(d)s
                """, conn, params={"d": st.session_state.dinas})

                df_sipd = pd.read_sql("""
                    SELECT kode_rekening, nama_rekening, saldo_akhir
                    FROM neraca_sipd
                    WHERE dinas = %(d)s
                """, conn, params={"d": st.session_state.dinas})

            # =========================
            # RENAME
            # =========================
            df_siap = df_siap.rename(columns={"saldo_akhir": "siap"})
            df_sipd = df_sipd.rename(columns={"saldo_akhir": "sipd"})

            # =========================
            # NUMERIC
            # =========================
            df_siap["siap"] = pd.to_numeric(df_siap["siap"], errors="coerce").fillna(0)
            df_sipd["sipd"] = pd.to_numeric(df_sipd["sipd"], errors="coerce").fillna(0)

            # =========================
            # MERGE
            # =========================
            df_merge = pd.merge(
                df_siap,
                df_sipd,
                on=["kode_rekening"],
                how="outer",
                suffixes=("_siap", "_sipd")
            )

            df_merge["nama_rekening"] = df_merge["nama_rekening_siap"].combine_first(
                df_merge["nama_rekening_sipd"]
            )

            df_merge["siap"] = df_merge["siap"].fillna(0)
            df_merge["sipd"] = df_merge["sipd"].fillna(0)

            # =========================
            # HITUNG SELISIH
            # =========================
            df_merge["selisih"] = df_merge["siap"] - df_merge["sipd"]

            st.session_state.df_merge = df_merge

            df_merge = df_merge[[
                "kode_rekening",
                "nama_rekening",
                "siap",
                "sipd",
                "selisih"
            ]]

            # =========================
            # DISPLAY
            # =========================
            def fmt(x):
                return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            df_display = df_merge.copy()
            df_display["siap"] = df_display["siap"].apply(fmt)
            df_display["sipd"] = df_display["sipd"].apply(fmt)
            df_display["selisih"] = df_display["selisih"].apply(fmt)

            st.subheader("📊 Perbandingan SIAP vs SIPD")
            st.dataframe(df_display, use_container_width=True)

            total_selisih = df_merge["selisih"].sum()

            st.markdown(
                f"""
                <div style="
                    margin-top:10px;
                    padding:10px;
                    border-radius:10px;
                    background-color:#f0f2f6;
                    text-align:right;
                    font-size:20px;
                    font-weight:bold;
                ">
                    Total Selisih : Rp {format_rupiah(total_selisih)}
                </div>
                """,
                unsafe_allow_html=True
            )

            # =========================
            # ✅ TOMBOL SIMPAN (PINDAH KE SINI)
            # =========================
            if st.button("💾 Simpan Hasil Perbandingan"):

                df_merge = st.session_state.df_merge

                with engine.begin() as conn:

                    conn.execute(text("""
                        DELETE FROM hasil_perbandingan
                        WHERE dinas = :d
                    """), {"d": st.session_state.dinas})

                    for _, row in df_merge.iterrows():

                        selisih = row["selisih"]

                        if selisih == 0:
                            continue

                        debit = selisih if selisih > 0 else 0
                        kredit = abs(selisih) if selisih < 0 else 0

                        conn.execute(text("""
                            INSERT INTO hasil_perbandingan (
                                nomor_bukti,
                                tanggal_bukti,
                                keterangan,
                                kode_bas,
                                uraian,
                                debit,
                                kredit,
                                keterangan_rinci,
                                dinas,
                                is_active,
                                created_at
                            )
                            VALUES (
                                :nomor, :tanggal, :ket,
                                :kode, :uraian,
                                :debit, :kredit,
                                :ket_rinci,
                                :dinas,
                                TRUE,
                                NOW()
                            )
                        """), {
                            "nomor": "JP Reviu Inspektorat/BBJ BLUD/2025",
                            "tanggal": "2025-12-31",
                            "ket": f"Jurnal Rinci BBJ BLUD {st.session_state.dinas}",
                            "kode": row["kode_rekening"],
                            "uraian": row["nama_rekening"],
                            "debit": debit,
                            "kredit": kredit,
                            "ket_rinci": "-",
                            "dinas": st.session_state.dinas
                        })

                st.success("✅ Hasil perbandingan berhasil disimpan")

            from io import BytesIO
            from openpyxl.styles import Alignment, Font, PatternFill

            # =========================
            # AMBIL DATA DARI DB
            # =========================
            with engine.begin() as conn:
                df_export = pd.read_sql("""
                    SELECT 
                        nomor_bukti AS "Nomor Bukti",
                        tanggal_bukti AS "Tanggal Bukti",
                        keterangan AS "Keterangan",
                        kode_bas AS "Kode BAS",
                        uraian AS "Uraian",
                        debit AS "Debit",
                        kredit AS "Kredit",
                        keterangan_rinci AS "Keterangan Rinci"
                    FROM hasil_perbandingan
                    WHERE dinas = %(d)s
                    ORDER BY kode_bas
                """, conn, params={"d": st.session_state.dinas})

            # =========================
            # KOSONGKAN KOLOM A-C (BARIS KE-2 DST)
            # =========================
            df_export_display = df_export.copy()

            cols_hide = ["Nomor Bukti", "Tanggal Bukti", "Keterangan"]

            for col in cols_hide:
                df_export_display.loc[1:, col] = ""

            # =========================
            # BUAT EXCEL
            # =========================
            output = BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export_display.to_excel(writer, index=False, sheet_name='jurnal')

                ws = writer.sheets['jurnal']

                # =========================
                # ALIGNMENT KOLOM A-C (ATAS)
                # =========================
                from openpyxl.styles import Alignment

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                    for cell in row:
                        cell.alignment = Alignment(vertical="top")

                # =========================
                # AUTO WIDTH (BIAR RAPI)
                # =========================
                column_widths = {
                    "A": 45,
                    "B": 18,
                    "C": 60,
                    "D": 18,
                    "E": 45,
                    "F": 18,
                    "G": 18,
                    "H": 25
                }

                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width

            output.seek(0)

            # =========================
            # DOWNLOAD BUTTON
            # =========================
            st.download_button(
                label="📥 Download Excel Penyesuaian",
                data=output,
                file_name=f"Jurnal_Penyesuaian_{st.session_state.dinas}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )